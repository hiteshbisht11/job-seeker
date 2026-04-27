"""Generate prioritized job-application Excel for Hitesh Bisht.

Sheets:
  1. Profile Summary       — resume distillation + fit-scoring rubric
  2. India - All Jobs      — every India role, fit-score sorted
  3. India - High Priority — fit >= 80
  4. India - Medium        — fit 65-79
  5. India - Stretch       — fit < 65
  6. Worldwide Remote      — global companies hiring India remote with high pay
  7. Application Tracker   — empty tracker pre-populated with company list

Run:  python src/generate_excel.py
"""

from __future__ import annotations

import os
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule

from jobs_data import JOBS, by_priority, all_sorted
from jobs_remote_global import JOBS_REMOTE_GLOBAL, by_priority as by_priority_remote
from profile import PROFILE


# ----- Styling helpers -------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1F4E78")
SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="595959")
BODY_FONT = Font(name="Calibri", size=10)
LINK_FONT = Font(name="Calibri", size=10, color="1F4E78", underline="single")

PRIORITY_FILL = {
    "High":    PatternFill("solid", fgColor="C6EFCE"),  # green
    "Medium":  PatternFill("solid", fgColor="FFEB9C"),  # amber
    "Stretch": PatternFill("solid", fgColor="FFC7CE"),  # red
}

THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row_idx: int, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row_idx].height = 32


def auto_size(ws, widths: list[int]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_link(ws, row: int, col: int, url: str, text: str | None = None):
    cell = ws.cell(row=row, column=col, value=text or url)
    cell.hyperlink = url
    cell.font = LINK_FONT
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = BORDER


# ----- Sheet builders --------------------------------------------------------

def build_profile_sheet(wb: Workbook):
    ws = wb.create_sheet("Profile Summary", 0)
    ws.sheet_view.showGridLines = False

    ws["A1"] = f"Job Application Plan — {PROFILE['name']}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    ws["A2"] = f"Generated {date.today().isoformat()}"
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:D2")

    rows = [
        ("Current Role",    PROFILE["current_role"]),
        ("Experience",      f"{PROFILE['experience_years']:.1f} years"),
        ("Education",       PROFILE["education"]),
        ("Location",        PROFILE["location"]),
    ]
    for i, (k, v) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

    ws.cell(row=9, column=1, value="Core Strengths").font = Font(bold=True, size=12, color="1F4E78")
    for i, s in enumerate(PROFILE["core_strengths"], start=10):
        ws.cell(row=i, column=1, value=f"  • {s}").alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)

    base = 10 + len(PROFILE["core_strengths"]) + 1
    ws.cell(row=base, column=1, value="Target Roles").font = Font(bold=True, size=12, color="1F4E78")
    for i, r in enumerate(PROFILE["target_roles"], start=base + 1):
        ws.cell(row=i, column=1, value=f"  • {r}").alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)

    base2 = base + 1 + len(PROFILE["target_roles"]) + 1
    ws.cell(row=base2, column=1, value="Fit-Score Rubric").font = Font(bold=True, size=12, color="1F4E78")
    for i, (k, v) in enumerate(PROFILE["fit_scoring_rubric"].items(), start=base2 + 1):
        ws.cell(row=i, column=1, value=k.replace("_", " ").title()).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

    legend_base = base2 + 1 + len(PROFILE["fit_scoring_rubric"]) + 2
    ws.cell(row=legend_base, column=1, value="Legend").font = Font(bold=True, size=12, color="1F4E78")
    legend = [
        ("High",    "Apply within first 7 days. Tailor resume bullets to the JD. Target: 15-20 applications."),
        ("Medium",  "Apply within 14 days. Standard resume + custom cover letter. Target: 25-30 applications."),
        ("Stretch", "Apply opportunistically. Use referrals where possible. Target: 5-10 applications."),
    ]
    for i, (k, v) in enumerate(legend, start=legend_base + 1):
        c = ws.cell(row=i, column=1, value=k)
        c.fill = PRIORITY_FILL[k]
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=2, value=v).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

    auto_size(ws, [22, 50, 30, 30])


# ----- Jobs sheet (India / generic) -----------------------------------------

INDIA_HEADERS = [
    "Priority", "Fit", "Company", "Role", "Category", "Location",
    "Experience", "Key Requirements", "Why Good Fit",
    "Careers Page", "LinkedIn Search", "Status", "Notes",
]
INDIA_WIDTHS = [10, 6, 24, 38, 22, 22, 12, 38, 50, 32, 32, 14, 28]


def write_jobs_sheet(wb: Workbook, name: str, jobs: list[dict], position: int | None = None):
    ws = wb.create_sheet(name, position) if position is not None else wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    # header
    for i, h in enumerate(INDIA_HEADERS, start=1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, 1, len(INDIA_HEADERS))

    for r, job in enumerate(jobs, start=2):
        ws.cell(row=r, column=1, value=job["priority"]).fill = PRIORITY_FILL[job["priority"]]
        ws.cell(row=r, column=2, value=job["fit_score"])
        ws.cell(row=r, column=3, value=job["company"]).font = Font(bold=True)
        ws.cell(row=r, column=4, value=job["role"])
        ws.cell(row=r, column=5, value=job["category"])
        ws.cell(row=r, column=6, value=job["location"])
        ws.cell(row=r, column=7, value=job["experience"])
        ws.cell(row=r, column=8, value=job["key_requirements"])
        ws.cell(row=r, column=9, value=job["why_fit"])
        write_link(ws, r, 10, job["primary_link"], "Careers Page →")
        write_link(ws, r, 11, job["search_link"], "LinkedIn Search →")
        ws.cell(row=r, column=12, value="To Apply")
        ws.cell(row=r, column=13, value="")

        for col in range(1, len(INDIA_HEADERS) + 1):
            cell = ws.cell(row=r, column=col)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = BORDER
            if col not in (10, 11):
                cell.font = cell.font.copy(name="Calibri", size=10)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 60

    # Conditional fill on Fit column (color scale)
    if len(jobs) > 0:
        last = len(jobs) + 1
        ws.conditional_formatting.add(
            f"B2:B{last}",
            ColorScaleRule(
                start_type="num", start_value=50, start_color="F8696B",
                mid_type="num",   mid_value=72,   mid_color="FFEB84",
                end_type="num",   end_value=95,   end_color="63BE7B",
            ),
        )

    auto_size(ws, INDIA_WIDTHS)


# ----- Worldwide remote sheet -----------------------------------------------

REMOTE_HEADERS = [
    "Priority", "Fit", "Company", "Role", "Category", "Location",
    "Experience", "Comp Band (INR)", "Key Requirements", "Why Good Fit",
    "Careers Page", "LinkedIn / Job Board", "Status", "Notes",
]
REMOTE_WIDTHS = [10, 6, 24, 38, 24, 24, 12, 18, 38, 50, 32, 32, 14, 28]


def write_remote_sheet(wb: Workbook, jobs: list[dict]):
    ws = wb.create_sheet("Worldwide Remote")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    for i, h in enumerate(REMOTE_HEADERS, start=1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, 1, len(REMOTE_HEADERS))

    for r, job in enumerate(jobs, start=2):
        ws.cell(row=r, column=1, value=job["priority"]).fill = PRIORITY_FILL[job["priority"]]
        ws.cell(row=r, column=2, value=job["fit_score"])
        ws.cell(row=r, column=3, value=job["company"]).font = Font(bold=True)
        ws.cell(row=r, column=4, value=job["role"])
        ws.cell(row=r, column=5, value=job["category"])
        ws.cell(row=r, column=6, value=job["location"])
        ws.cell(row=r, column=7, value=job["experience"])
        ws.cell(row=r, column=8, value=job["comp_band_inr"])
        ws.cell(row=r, column=9, value=job["key_requirements"])
        ws.cell(row=r, column=10, value=job["why_fit"])
        write_link(ws, r, 11, job["primary_link"], "Careers Page →")
        write_link(ws, r, 12, job["search_link"], "Job Board →")
        ws.cell(row=r, column=13, value="To Apply")
        ws.cell(row=r, column=14, value="")

        for col in range(1, len(REMOTE_HEADERS) + 1):
            cell = ws.cell(row=r, column=col)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = BORDER
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 64

    if len(jobs) > 0:
        last = len(jobs) + 1
        ws.conditional_formatting.add(
            f"B2:B{last}",
            ColorScaleRule(
                start_type="num", start_value=50, start_color="F8696B",
                mid_type="num",   mid_value=72,   mid_color="FFEB84",
                end_type="num",   end_value=95,   end_color="63BE7B",
            ),
        )

    auto_size(ws, REMOTE_WIDTHS)


# ----- Application tracker --------------------------------------------------

TRACKER_HEADERS = [
    "Company", "Role", "Source Sheet", "Date Applied", "Resume Version",
    "Referral?", "Recruiter Contact", "Stage", "Next Step", "Notes",
]
TRACKER_WIDTHS = [22, 38, 22, 14, 18, 12, 24, 18, 28, 32]
STAGES = ["Not Started", "Applied", "Recruiter Screen", "Tech Round 1",
          "Tech Round 2", "System Design", "Hiring Manager", "Offer", "Rejected"]


def build_tracker_sheet(wb: Workbook):
    ws = wb.create_sheet("Application Tracker")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    for i, h in enumerate(TRACKER_HEADERS, start=1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, 1, len(TRACKER_HEADERS))

    rows: list[tuple[str, str, str]] = []
    for j in all_sorted():
        rows.append((j["company"], j["role"], "India"))
    for j in JOBS_REMOTE_GLOBAL:
        rows.append((j["company"], j["role"], "Worldwide Remote"))

    for r, (company, role, src) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=company).font = Font(bold=True)
        ws.cell(row=r, column=2, value=role)
        ws.cell(row=r, column=3, value=src)
        ws.cell(row=r, column=8, value="Not Started")

        for col in range(1, len(TRACKER_HEADERS) + 1):
            cell = ws.cell(row=r, column=col)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = BORDER
        ws.row_dimensions[r].height = 28

    auto_size(ws, TRACKER_WIDTHS)


# ----- Main ------------------------------------------------------------------

def main():
    wb = Workbook()
    # remove default empty sheet
    default = wb.active
    wb.remove(default)

    build_profile_sheet(wb)
    write_jobs_sheet(wb, "India - All Jobs",      all_sorted())
    write_jobs_sheet(wb, "India - High Priority", by_priority("High"))
    write_jobs_sheet(wb, "India - Medium",        by_priority("Medium"))
    write_jobs_sheet(wb, "India - Stretch",       by_priority("Stretch"))
    write_remote_sheet(wb, sorted(JOBS_REMOTE_GLOBAL, key=lambda j: j["fit_score"], reverse=True))
    build_tracker_sheet(wb)

    out_dir = os.path.join(os.path.dirname(__file__), "..", "output")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.abspath(os.path.join(out_dir, "hitesh_job_targets.xlsx"))
    wb.save(out_path)
    print(f"Wrote {out_path}")
    print(f"  India roles:     {len(JOBS)}")
    print(f"    High:          {len(by_priority('High'))}")
    print(f"    Medium:        {len(by_priority('Medium'))}")
    print(f"    Stretch:       {len(by_priority('Stretch'))}")
    print(f"  Remote (global): {len(JOBS_REMOTE_GLOBAL)}")


if __name__ == "__main__":
    main()
